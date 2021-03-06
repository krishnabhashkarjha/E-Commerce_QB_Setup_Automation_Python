'''
@ Author - Yogesh Pawar
@ Creation date - 02/10/2020
@ Description - SVC Utility
'''

#1st parameter = requird parameter sheet
#2nd parameter = excel for data process
#3rd parameter = which column value end user looking for

import openpyxl

def svc_utility_excel_process(input_file_path,process_file_path,process_file_column_number_to_get_value):
    v_input_wb = openpyxl.load_workbook(input_file_path)
    v_input_sheet = v_input_wb.get_sheet_by_name("Input")
    total_rows_input=v_input_sheet.max_row

    v_processed_wb = openpyxl.load_workbook(process_file_path)
    v_processed_sheet = v_processed_wb.get_sheet_by_name("processed")
    total_rows_processed=v_processed_sheet.max_row
    total_columns_processed=v_processed_sheet.max_column

    for i in range(int(total_rows_processed)):
        for j in range(int(total_rows_input)):
            for k in range(int(total_columns_processed)):
                if str(v_input_sheet.cell(row=(j+1), column=1).value)==str(v_processed_sheet.cell(row=(i+1), column=(k+1)).value):
                    v_input_sheet.cell(row=(j+1), column=2).value=v_processed_sheet.cell(row=(i+1), column=int(process_file_column_number_to_get_value)).value
                    v_input_wb.save(input_file_path)
                    print("Done")

svc_utility_excel_process("C:/Users/yogesh.pawar/Desktop/SVC_Shrort_automation/Input.xlsx","C:/Users/yogesh.pawar/Desktop/SVC_Shrort_automation/processed.xlsx","1")
